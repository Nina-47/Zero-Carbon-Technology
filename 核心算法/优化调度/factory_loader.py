# -*- coding: utf-8 -*-
"""
工厂数据加载（通用化）：读任意厂的负荷 + 中山天气辐射。

负荷：从「逐时数据」sheet 的「实际用电量」列读（MWh→kW），
      含光伏自发自用（A 有光伏、B/C 无光伏时该列 = 电网电量）。
辐射：从天气数据「长表」读逐时辐射 W/m²，按日期对齐负荷。

这是「容量通用化」的数据入口：任何厂只要负荷表列布局一致
（0日期 | 1-24电网 | 25-48光伏 | 49-72实际用电量），就能读进来。
光伏出力不在这里读——它由装机 + 辐射经 pv_output_model 计算，
因此 B/C/D 无光伏历史数据也能算。
"""

import numpy as np
import openpyxl
from datetime import datetime

import config


# 厂名 → 负荷表路径
FACTORY_XLSX = {
    "A": config.DATA_SOURCE_XLSX,
    "B": config.DATA_SOURCE_B,
    "C": config.DATA_SOURCE_C,
}


def _parse_date(v):
    return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()


def load_load(factory):
    """读指定厂的逐时负荷（实际用电量列，MWh→kW）。

    返回 (load_arr (N,24) kW, dates list, day_num)。
    """
    path = FACTORY_XLSX[factory.upper()]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["逐时数据"]
    loads, dates = [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        ld = [0.0 if v is None else float(v) for v in row[49:73]]  # 实际用电量 MWh
        if len(ld) == 24:
            dates.append(_parse_date(row[0]))
            loads.append(ld)
    wb.close()
    return np.array(loads) * 1000.0, dates, len(loads)


def load_radiation(dates):
    """读中山逐时辐射 W/m²，按给定日期列表对齐（天气「长表」）。

    返回 rad_arr (len(dates), 24) W/m²；某日缺失时该日全 0。
    """
    wb = openpyxl.load_workbook(config.WEATHER_XLSX, read_only=True, data_only=True)
    ws = wb["长表"]
    header = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header = row
        break
    i_date = header.index("date")
    i_hour = header.index("hour")
    i_rad = header.index("辐射_Wm2")
    tmp = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_date] is None:
            continue
        d = _parse_date(row[i_date])
        h = int(row[i_hour])
        v = 0.0 if row[i_rad] is None else float(row[i_rad])
        tmp.setdefault(d, [0.0] * 24)[h] = v
    wb.close()
    return np.array([tmp.get(d, [0.0] * 24) for d in dates])


def load_factory(factory):
    """通用入口：读负荷 + 对齐辐射。

    返回 dict {load: (N,24) kW, rad: (N,24) W/m², dates: list, day_num: int}。
    """
    load_arr, dates, day_num = load_load(factory)
    rad_arr = load_radiation(dates)
    return {
        "load": load_arr,
        "rad": rad_arr,
        "dates": dates,
        "day_num": day_num,
    }
