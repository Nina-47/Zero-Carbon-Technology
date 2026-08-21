# -*- coding: utf-8 -*-
"""
数据整理：把 A/B/C 三家污水厂拆成三个独立表格（负荷 + 光伏 + 实际用电量）

输入：
  - 02_原始数据/负荷数据/三污水厂逐时电量.xlsx  （A/B/C 三家电网净购电，逐时 MWh）
  - 02_原始数据/光伏数据/中嘉光伏数据.xlsx       （仅 A=中嘉 有光伏，逐时 kWh + 日自发自用拆分）

输出（写到 02_原始数据/负荷数据/）：
  - A_中嘉污水厂.xlsx        （含光伏）
  - B_中山市污水处理.xlsx    （无光伏）
  - C_珍家山污水厂.xlsx      （无光伏）
  每个文件 3 个 sheet：说明 / 逐时数据 / 日汇总

口径说明（重要）：
  电网电量   = 从电网买的电（净购电，电表记的）          —— 单位 MWh
  光伏发电   = 三个场站（一期+二期+车棚）自己发的电       —— 原数据 kWh，此处统一 MWh
  自发自用   = 光伏发电里被污水厂自己消纳的部分
  余电上网   = 光伏发电里卖给电网的部分（反向）
  实际用电量 = 电网电量 + 自发自用  （污水厂真正消耗的总负荷）

  关系：光伏发电 = 自发自用 + 余电上网
        实际用电量 = 电网电量 + 自发自用

  逐时自发自用拆分：原始数据里「自发自用/余电上网」只有天级（每天一行），没有逐时。
  因此逐时的自发自用按「当日自发自用率 × 该小时光伏发电」近似分配：
        当日自发自用率 = 日自发自用 / 日光伏发电
        逐时自发自用   = 逐时光伏发电 × 当日自发自用率
  这是能量核算里标准的「按出力曲线分摊」，误差很小（余电上网整体只占 ~5%）。

数据范围：2025-07-01 ~ 2026-07-31（396 天，连续）
  （2025-04 有 30 天电网数据，但缺 5/6 月且当时没光伏，本表不纳入）
"""
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"C:/Users/xiaoY/Desktop/零碳科技"
LOAD_FILE = f"{BASE}/02_原始数据/负荷数据/三污水厂逐时电量.xlsx"
PV_FILE = f"{BASE}/02_原始数据/光伏数据/中嘉光伏数据.xlsx"
OUT_DIR = f"{BASE}/02_原始数据/负荷数据"

START = datetime.date(2025, 7, 1)
END = datetime.date(2026, 7, 31)
DATES = []
d = START
while d <= END:
    DATES.append(d)
    d += datetime.timedelta(days=1)
print(f"目标区间 {START} ~ {END}，共 {len(DATES)} 天")


def load_grid_hourly(sheet_name):
    """读取三污水厂逐时电量.xlsx 某公司 sheet，返回 {date_str: [24小时MWh]}"""
    wb = openpyxl.load_workbook(LOAD_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    out = {}
    for r in ws.iter_rows(values_only=True):
        if r[0] is None or not isinstance(r[0], str):
            continue
        date = r[0].strip()
        if date == "日期":
            continue
        # 合计电量 r[1]；逐时 r[2:26]
        hours = [float(x) if isinstance(x, (int, float)) else 0.0 for x in r[2:26]]
        if len(hours) < 24:
            hours = hours + [0.0] * (24 - len(hours))
        out[date] = hours[:24]
    wb.close()
    return out


def load_pv_hourly(station_sheet):
    """读取光伏逐时 sheet，返回 {date_str: [24小时 kWh]}"""
    wb = openpyxl.load_workbook(PV_FILE, read_only=True, data_only=True)
    ws = wb[station_sheet]
    out = {}
    for r in ws.iter_rows(values_only=True):
        t = r[3]  # 对应时间
        if t is None or not isinstance(t, datetime.datetime):
            continue
        date = t.strftime("%Y-%m-%d")
        # 总发电量 r[6]，逐时 0点~23点 在 r[7:31]
        hours = [float(x) if isinstance(x, (int, float)) else 0.0 for x in r[7:31]]
        if len(hours) < 24:
            hours = hours + [0.0] * (24 - len(hours))
        out[date] = hours[:24]
    wb.close()
    return out


def load_pv_daily():
    """读取数据统计分析表，返回 {date_str: dict(光伏发电kWh, 自发自用kWh, 余电上网kWh, 电网买电MWh)}"""
    wb = openpyxl.load_workbook(PV_FILE, read_only=True, data_only=True)
    ws = wb["数据统计分析表"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        t = r[0]
        if t is None or not isinstance(t, datetime.datetime):
            continue
        date = t.strftime("%Y-%m-%d")
        # [1]光伏总发电 [4]车棚发电 [5]电网买电MWh [9]一期自用 [10]一期反向 [13]二期自用 [14]二期反向
        pv_total = float(r[1]) if isinstance(r[1], (int, float)) else 0.0
        carport = float(r[4]) if isinstance(r[4], (int, float)) else 0.0
        grid = float(r[5]) if isinstance(r[5], (int, float)) else 0.0
        yq_self = float(r[9]) if isinstance(r[9], (int, float)) else 0.0
        yq_rev = float(r[10]) if isinstance(r[10], (int, float)) else 0.0
        eq_self = float(r[13]) if isinstance(r[13], (int, float)) else 0.0
        eq_rev = float(r[14]) if isinstance(r[14], (int, float)) else 0.0
        # 自发自用 = 一期自用 + 二期自用 + 车棚发电（车棚全部自用）
        self_use = yq_self + eq_self + carport
        # 余电上网 = 一期反向 + 二期反向
        reverse = yq_rev + eq_rev
        out[date] = dict(pv=pv_total, self_use=self_use, reverse=reverse, grid=grid)
    wb.close()
    return out


# ---------- 读取 ----------
grid = {c: load_grid_hourly(c) for c in ["A公司", "B公司", "C公司"]}
pv_hourly = {
    "一期": load_pv_hourly("中嘉污水处理厂8MWp光伏电站(一期)"),
    "二期": load_pv_hourly("中嘉污水处理厂8MWp光伏电站(二期)"),
    "车棚": load_pv_hourly("中嘉污水处理厂光伏车棚项目"),
}
pv_daily = load_pv_daily()

print("电网逐时：A=%d 天, B=%d, C=%d" % (len(grid["A公司"]), len(grid["B公司"]), len(grid["C公司"])))
print("光伏逐时：一期=%d, 二期=%d, 车棚=%d" % (len(pv_hourly["一期"]), len(pv_hourly["二期"]), len(pv_hourly["车棚"])))
print("光伏日汇总：%d 天" % len(pv_daily))


# ---------- 组装 ----------
def build_company(name, ggrid, has_pv):
    """返回 (逐时矩阵, 日汇总矩阵)。矩阵每行 = [date, ...]"""
    hourly_rows = []  # [date, grid*24, pv*24, actual*24]
    daily_rows = []   # [date, pv, self_use, reverse, grid, actual]
    mismatch = 0
    missing_grid = 0
    for date in DATES:
        ds = date.strftime("%Y-%m-%d")
        gh = ggrid.get(ds)
        if gh is None:
            gh = [0.0] * 24
            missing_grid += 1

        if has_pv:
            # 逐时光伏 = 一期+二期+车棚 (kWh -> MWh)
            ph = [0.0] * 24
            for k in ("一期", "二期", "车棚"):
                v = pv_hourly[k].get(ds, [0.0] * 24)
                for i in range(24):
                    ph[i] += v[i] / 1000.0
            # 日汇总数据
            dd = pv_daily.get(ds)
            if dd is not None:
                pv_mwh = dd["pv"] / 1000.0
                self_mwh = max(0.0, dd["self_use"] / 1000.0)  # 负值（调试期）截为 0
                rev_mwh = dd["reverse"] / 1000.0
                grid_mwh = dd["grid"]
            else:
                pv_mwh = sum(ph)
                self_mwh = pv_mwh  # 无拆分数据时近似全部自用
                rev_mwh = 0.0
                grid_mwh = sum(gh)
            # 校验：光伏发电 ≈ 自发自用 + 余电上网
            if abs(pv_mwh - (self_mwh + rev_mwh)) > 0.05:
                mismatch += 1
            # 当日自发自用率
            rate = self_mwh / pv_mwh if pv_mwh > 0 else 0.0
            rate = max(0.0, min(1.0, rate))
            # 逐时实际用电 = 电网 + 逐时光伏 × rate
            ah = [gh[i] + ph[i] * rate for i in range(24)]
            actual_mwh = grid_mwh + self_mwh
            hourly_rows.append([date] + gh + ph + ah)
            daily_rows.append([date, pv_mwh, self_mwh, rev_mwh, grid_mwh, actual_mwh])
        else:
            ph = [0.0] * 24
            ah = gh  # 无光伏：实际用电量 = 电网电量
            grid_mwh = sum(gh)
            hourly_rows.append([date] + gh + ph + ah)
            daily_rows.append([date, 0.0, 0.0, 0.0, grid_mwh, grid_mwh])
    print(f"[{name}] 缺电网日数={missing_grid}, 光伏校验不符日数={mismatch}")
    return hourly_rows, daily_rows


A_h, A_d = build_company("A公司", grid["A公司"], has_pv=True)
B_h, B_d = build_company("B公司", grid["B公司"], has_pv=False)
C_h, C_d = build_company("C公司", grid["C公司"], has_pv=False)


# ---------- 写出 ----------
HOURS_CN = ["0时", "1时", "2时", "3时", "4时", "5时", "6时", "7时", "8时", "9时",
            "10时", "11时", "12时", "13时", "14时", "15时", "16时", "17时", "18时",
            "19时", "20时", "21时", "22时", "23时"]

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
TYPE_FILL = PatternFill("solid", fgColor="BDD7EE")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def write_company(fname, name, has_pv, hourly, daily):
    wb = openpyxl.Workbook()

    # ---- 说明 sheet ----
    ws0 = wb.active
    ws0.title = "说明"
    lines = [
        [f"{name} 三类数据表（负荷 / 光伏 / 实际用电量）"],
        [""],
        ["单位：全部为 MWh（兆瓦时）"],
        [""],
        ["三类数据含义："],
        ["  电网电量   = 从电网买的电（净购电，电表记的），就是污水厂每个月的电费账单口径"],
        ["  光伏发电   = 场站自己发的电（A=中嘉有光伏：一期+二期+车棚；B/C 无光伏，均为 0）"],
        ["  实际用电量 = 污水厂真正消耗的总负荷 = 电网电量 + 光伏自发自用"],
        [""],
        ["光伏相关（仅 A 有）："],
        ["  自发自用 = 光伏发电里被污水厂自己消纳的部分（不花钱）"],
        ["  余电上网 = 光伏发电里卖给电网的部分（反向，有收益）"],
        ["  关系：光伏发电 = 自发自用 + 余电上网"],
        [""],
        ["数据范围：2025-07-01 ~ 2026-07-31（396 天，连续）"],
        ["数据来源：三污水厂逐时电量.xlsx + 中嘉光伏数据.xlsx"],
        [""],
        ["注意：逐时表里「实际用电量」的自发自用部分按当日自发自用率分摊（原始只有天级拆分），"],
        ["     日汇总表里的自发自用/余电上网为原始精确值。"],
    ]
    for row in lines:
        ws0.append(row)
    ws0.column_dimensions["A"].width = 60

    # ---- 逐时数据 sheet ----
    ws1 = wb.create_sheet("逐时数据")
    # 第一行：类型（合并单元格）
    ws1.cell(1, 1, "日期")
    ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws1.cell(1, 2, "电网电量(MWh)")
    ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=25)
    ws1.cell(1, 26, "光伏发电(MWh)")
    ws1.merge_cells(start_row=1, start_column=26, end_row=1, end_column=49)
    ws1.cell(1, 50, "实际用电量(MWh)")
    ws1.merge_cells(start_row=1, start_column=50, end_row=1, end_column=73)
    # 第二行：小时（A1 已与 B1 合并，故从第 2 列开始写）
    for i, h in enumerate(HOURS_CN):
        ws1.cell(2, 2 + i, h)
        ws1.cell(2, 26 + i, h)
        ws1.cell(2, 50 + i, h)
    # 样式
    for c in range(1, 74):
        ws1.cell(1, c).fill = TYPE_FILL
        ws1.cell(1, c).font = BOLD
        ws1.cell(1, c).alignment = CENTER
        ws1.cell(2, c).fill = HEADER_FILL
        ws1.cell(2, c).font = BOLD
        ws1.cell(2, c).alignment = CENTER
    # 数据
    for row in hourly:
        date = row[0].strftime("%Y-%m-%d")
        vals = [round(x, 4) for x in row[1:]]
        ws1.append([date] + vals)
    ws1.freeze_panes = "B3"

    # ---- 日汇总 sheet ----
    ws2 = wb.create_sheet("日汇总")
    cols = ["日期", "光伏发电(MWh)", "自发自用(MWh)", "余电上网(MWh)", "电网电量(MWh)", "实际用电量(MWh)"]
    ws2.append(cols)
    for c in range(1, 7):
        ws2.cell(1, c).fill = HEADER_FILL
        ws2.cell(1, c).font = BOLD
        ws2.cell(1, c).alignment = CENTER
    for row in daily:
        date = row[0].strftime("%Y-%m-%d")
        vals = [round(x, 4) for x in row[1:]]
        ws2.append([date] + vals)
    for c in range(1, 7):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15
    ws2.freeze_panes = "A2"

    out = f"{OUT_DIR}/{fname}"
    wb.save(out)
    print("已写出:", out)


write_company("A_中嘉污水厂.xlsx", "A 中嘉污水厂", True, A_h, A_d)
write_company("B_中山市污水处理.xlsx", "B 中山市污水处理", False, B_h, B_d)
write_company("C_珍家山污水厂.xlsx", "C 珍家山污水厂", False, C_h, C_d)

print("\n完成。三家分表已生成。")
