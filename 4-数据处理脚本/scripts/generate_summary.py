# -*- coding: utf-8 -*-
"""
生成最终成果汇总表（Excel）——固化模块二调度 + 模块三碳核算全年结果。

输出：3-分析报告文档/零碳项目_全年成果汇总_20260819.xlsx
包含 4 张表：
  1. 总体结论    核心 KPI 一览
  2. 模块二调度  全年电量/成本/储能/双模式汇总
  3. 模块三碳核算 三账（企业清单/项目减排/LCA）结果
  4. 数据口径    关键参数与数据来源说明
"""

import sys
import os
import json
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# 路径
BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # 公用/公用（本脚本位于 4-数据处理脚本/scripts/）
MODULE2_JSON = os.path.join(BASE, "模块二", "智能调度", "output", "yearly_dispatch.json")
MODULE3_JSON = os.path.join(BASE, "模块三", "code", "carbon_results.json")
OUT = os.path.join(BASE, "3-分析报告文档", "零碳项目_全年成果汇总_20260819.xlsx")


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_module2_summary():
    d = load_json(MODULE2_JSON)
    days = d["days"]
    load = sum(x["summary"]["load_kWh"] for x in days)
    pv = sum(x["summary"]["pv_kWh"] for x in days)
    buy = sum(x["summary"]["buy_kWh"] for x in days)
    export = sum(x["summary"]["export_kWh"] for x in days)
    cost = sum(x["summary"]["cost_yuan"] for x in days)
    chg = sum(x["summary"]["chg_kWh"] for x in days)
    dis = sum(x["summary"]["dis_kWh"] for x in days)
    chg_pv = sum(x["summary"]["chg_pv_kWh"] for x in days)
    dis_replace = sum(x["summary"]["dis_replace_kWh"] for x in days)
    safe = sum(1 for x in days if x["mode"] == "safe")
    shock = sum(1 for x in days if x["is_shock"])
    green = (load - buy) / load * 100 if load else 0
    return {
        "days": len(days),
        "start": d["meta"]["start_date"],
        "end": d["meta"]["end_date"],
        "load_万kWh": round(load / 1e4, 1),
        "pv_万kWh": round(pv / 1e4, 1),
        "buy_万kWh": round(buy / 1e4, 1),
        "export_万kWh": round(export / 1e4, 1),
        "cost_万元": round(cost / 1e4, 0),
        "chg_万kWh": round(chg / 1e4, 1),
        "dis_万kWh": round(dis / 1e4, 1),
        "chg_pv_万kWh": round(chg_pv / 1e4, 1),
        "dis_replace_万kWh": round(dis_replace / 1e4, 1),
        "green_pct": round(green, 2),
        "safe_days": safe,
        "shock_days": shock,
        "eco_days": len(days) - safe,
    }


def build_module3_summary():
    r = load_json(MODULE3_JSON)
    res = r["results"]
    s1 = res["scope1"]
    s2 = res["scope2"]
    pr = res["project"]
    lc = res["lca"]
    return {
        "scope1_总_t": round(s1["scope1_kg"] / 1e3, 1),
        "ch4_t": round(s1["ch4_co2e_kg"] / 1e3, 1),
        "n2o_t": round(s1["n2o_co2e_kg"] / 1e3, 1),
        "scope2_总_t": round(s2["scope2_kg"] / 1e3, 1),
        "购电_万kWh": round(s2["e_grid_kwh"] / 1e4, 1),
        "厂区总排放_t": round(res["total_plant_kg"] / 1e3, 1),
        "BE_t": round(pr["BE_kg"] / 1e3, 1),
        "PE_t": round(pr["PE_kg"] / 1e3, 1),
        "LE_t": round(pr["LE_kg"] / 1e3, 1),
        "ER_t": round(pr["ER_kg"] / 1e3, 1),
        "隐含碳_总_t": round(lc["embodied_total_kg"] / 1e3, 1),
        "隐含碳_净_t": round(lc["embodied_net_kg"] / 1e3, 1),
        "隐含碳_年均分摊_t": round(lc["embodied_annual_kg"] / 1e3, 1),
        "净生命周期年减排_t": round(lc["er_net_yearly_kg"] / 1e3, 1),
        "寿命可放电_万kWh": round(lc["e_life_kwh"] / 1e4, 1),
    }


def write_sheet(ws, title, rows, header_fill="4472C4"):
    ws.append([title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=header_fill)
    for r in rows:
        ws.append(r)
    # 表头加粗
    for cell in ws[2]:
        cell.font = Font(bold=True)


def main():
    m2 = build_module2_summary()
    m3 = build_module3_summary()

    wb = openpyxl.Workbook()

    # Sheet1 总体结论
    ws1 = wb.active
    ws1.title = "总体结论"
    rows1 = [
        ["指标", "数值", "单位", "说明"],
        ["厂区年度总碳排放", m3["厂区总排放_t"], "吨 CO2e", "范围一+范围二"],
        ["  范围一（CH4+N2O 工艺）", m3["scope1_总_t"], "吨 CO2e", "污水处理过程直接排放"],
        ["  范围二（外购电力）", m3["scope2_总_t"], "吨 CO2", "净购电 × 0.4419"],
        ["运行减排量 ER", m3["ER_t"], "吨 CO2e/年", "基准 - 项目 - 泄漏"],
        ["净生命周期减排", m3["净生命周期年减排_t"], "吨 CO2e/年", "运行减排 - 隐含碳分摊"],
        ["光伏装机（三站合计×2规划）", "12.8", "MW", "6.4MW真实×2放大"],
        ["光伏年发电量", m2["pv_万kWh"], "万 kWh", "三站合计"],
        ["光伏绿电占比", m2["green_pct"], "%", "自发自用/总负荷"],
        ["储能配置", "4/2", "MWh/MW", "改进版"],
        ["年节省电费（现货价结算）", m2["cost_万元"], "万元", "真实节点现货价"],
    ]
    write_sheet(ws1, "零碳项目 · 全年成果总览", rows1, "2E7D32")

    # Sheet2 模块二调度
    ws2 = wb.create_sheet("模块二调度")
    rows2 = [
        ["指标", "数值", "单位"],
        ["调度天数", m2["days"], "天"],
        ["日期范围", f"{m2['start']} ~ {m2['end']}", ""],
        ["全年总负荷", m2["load_万kWh"], "万 kWh"],
        ["全年光伏发电", m2["pv_万kWh"], "万 kWh"],
        ["全年外购电", m2["buy_万kWh"], "万 kWh"],
        ["全年余电上网", m2["export_万kWh"], "万 kWh"],
        ["全年购电成本（现货价）", m2["cost_万元"], "万元"],
        ["储能年充电量", m2["chg_万kWh"], "万 kWh"],
        ["储能年放电量", m2["dis_万kWh"], "万 kWh"],
        ["  其中光伏充电", m2["chg_pv_万kWh"], "万 kWh"],
        ["  放电替代购电", m2["dis_replace_万kWh"], "万 kWh"],
        ["安全模式天数", m2["safe_days"], "天"],
        ["节能模式天数", m2["eco_days"], "天"],
        ["冲击负荷豁免天数", m2["shock_days"], "天"],
        ["绿电占比", m2["green_pct"], "%"],
    ]
    write_sheet(ws2, "模块二 · 智能调度全年汇总", rows2, "E67E22")

    # Sheet3 模块三碳核算
    ws3 = wb.create_sheet("模块三碳核算")
    rows3 = [
        ["账本", "指标", "数值(吨)", "说明"],
        ["企业清单账", "范围一 CH4", m3["ch4_t"], "厌氧产甲烷"],
        ["", "范围一 N2O", m3["n2o_t"], "脱氮过程"],
        ["", "范围一合计", m3["scope1_总_t"], "CH4+N2O"],
        ["", "范围二外购电", m3["scope2_总_t"], f"购电{m3['购电_万kWh']}万kWh×0.4419"],
        ["", "厂区总排放", m3["厂区总排放_t"], "范围一+二"],
        ["项目减排账", "基准排放 BE", m3["BE_t"], "无储能全电网"],
        ["", "项目排放 PE", m3["PE_t"], "光储柔优化后"],
        ["", "泄漏 LE", m3["LE_t"], "辅助电耗"],
        ["", "运行减排 ER", m3["ER_t"], "BE-PE-LE"],
        ["LCA账", "储能隐含碳总量", m3["隐含碳_总_t"], "4MWh×70kg/kWh"],
        ["", "扣回收后", m3["隐含碳_净_t"], "回收抵减10%"],
        ["", "年均分摊", m3["隐含碳_年均分摊_t"], "按12年"],
        ["", "净生命周期年减排", m3["净生命周期年减排_t"], "ER-分摊"],
    ]
    write_sheet(ws3, "模块三 · 碳核算三账结果", rows3, "C0392B")

    # Sheet4 数据口径
    ws4 = wb.create_sheet("数据口径")
    rows4 = [
        ["参数", "取值", "来源"],
        ["广东电力排放因子 EF_report", "0.4419 kgCO2/kWh", "生态环境部2023省级因子公告"],
        ["时变碳因子方法", "节点实时电价代理", "现货价格→碳强度映射"],
        ["实时电价均值", "0.333 元/kWh", "广东电力现货节点实时电价"],
        ["CH4参数 B0/MCF", "0.25 / 0.05", "IPCC 2006 + AAO工艺"],
        ["N2O因子", "0.016 kgN2O-N/kgN", "Li et al. 2023 国内AAO"],
        ["进水/出水 COD", "150 / 10 mg/L", "广东统计中值 + SCADA照片"],
        ["进水/出水 TN", "30 / 7 mg/L", "同上"],
        ["日处理水量", "18.4 万m3/d", "电耗反算"],
        ["光伏装机", "6.4MW真实 ×2=12.8MW", "30公顷面积约束规划"],
        ["储能", "4MWh/2MW, η=90%", "宁德时代EnerC改进版"],
        ["LFP隐含碳", "70 kgCO2e/kWh", "行业LCA荟萃值"],
        ["储能循环寿命", "6000次@80%DoD", "宁德时代液冷白皮书"],
    ]
    write_sheet(ws4, "数据口径与来源", rows4, "6C3483")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"汇总表已生成: {OUT}")


if __name__ == "__main__":
    main()
