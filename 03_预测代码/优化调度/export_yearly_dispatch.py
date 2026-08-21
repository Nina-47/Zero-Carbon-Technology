# -*- coding: utf-8 -*-
"""
逐日调度导出：过去一年（按真实数据天数）每天的详细调度建议 → JSON。

与 validate_optimizer.backtest_full_year 使用完全相同的逐日滚动链路
（双模式状态机 + daily_opt_dispatch），但把每一天的逐时结果和汇总都存下来，
统一导出 JSON，供碳核算模块三、网页大屏等接口消费。

用法：
  python export_yearly_dispatch.py [输出json路径]

输出 JSON 结构：
  {
    "meta": {储能/光伏/电价等配置快照},
    "days": [
      {
        "date": "2025-07-01",
        "idx": 0,
        "mode": "eco"/"safe",
        "is_shock": false,
        "flex_min": 0.65,
        "summary": {日总负荷/光伏/购电/余电上网/绿电占比/成本/储能充放...},
        "hourly": [ {h, load_kW, pv_kW, pch_kW, pdis_kW, pgrid_kW, flex_kW, flex_down_kW, soc}, ... 24项 ]
      }, ...
    ]
  }
"""

import sys
import os
import json
import numpy as np
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config
from storage_optimizer import daily_opt_dispatch
from mode_controller import ModeController
from validate_optimizer import load_data_source


def build_yearly_dispatch():
    """逐日滚动跑全年，返回 {meta, days} 结构。"""
    pv_arr, load_arr, day_num = load_data_source()

    # 起始日期：A 分表「逐时数据」首行日期（load_data_source 内部已对齐）
    from openpyxl import load_workbook
    wb = load_workbook(config.DATA_SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["逐时数据"]
    first_date_cell = None
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        first_date_cell = row[0]
        break
    wb.close()
    start_date = str(first_date_cell)[:10] if first_date_cell is not None else \
        (datetime(2025, 7, 1) + timedelta(days=0)).strftime("%Y-%m-%d")

    # 双模式状态机
    mc = ModeController(
        flex_min=config.FLEX_MIN, flex_min_safe=config.FLEX_MIN_SAFE,
        flex_max=config.FLEX_MAX,
        flex_down_cont_h=config.FLEX_DOWN_CONT_H,
        load_shock_ratio=config.LOAD_SHOCK_RATIO,
        hysteresis_days=config.HYSTERESIS_DAYS,
        flex_energy_ratio=config.FLEX_ENERGY_RATIO,
    )

    # 真实现货价（元/kWh）；缺失则回退固定峰谷价
    realtime_price = config.load_realtime_price()
    fallback_price = np.array([config.tou_price(t) for t in range(24)])
    price_used = "现货" if realtime_price else "峰谷目录价"

    soc_now = config.SOC_INIT

    days = []
    fail_days = 0

    for d in range(day_num):
        pv = pv_arr[d]
        load = load_arr[d]

        # 当日价格：优先真实现货价，否则固定峰谷价
        date_str = _date_str(start_date, d)
        price = realtime_price.get(datetime.strptime(date_str, "%Y-%m-%d").date(),
                                   fallback_price) if realtime_price else fallback_price

        flex_min_override, flex_energy_ratio, mode, is_shock = mc.decide(load)

        res = daily_opt_dispatch(
            pv_kw=pv, load_kw=load,
            E_max=config.E_BAT_MAX, P_max=config.P_BAT_MAX,
            eta_ch=config.ETA_CH, eta_dis=config.ETA_DIS,
            soc_min=config.SOC_MIN, soc_max=config.SOC_MAX,
            soc0=soc_now, price=price, w_cost=config.W_COST, w_green=config.W_GREEN,
            flex_min=config.FLEX_MIN, flex_max=config.FLEX_MAX,
            flex_penalty=config.FLEX_PENALTY, flex_enabled=config.FLEX_ENABLED,
            flex_min_override=flex_min_override,
            flex_dev_cost=config.FLEX_DEV_COST,
            flex_energy_ratio=flex_energy_ratio,
        )

        if res.get("soc") is None:
            fail_days += 1
            # 无解那天输出占位，保持日期连续，便于接口消费
            days.append({
                "date": start_date if d == 0 else "",
                "idx": d,
                "mode": mode,
                "is_shock": bool(is_shock),
                "flex_min": round(flex_min_override, 3),
                "status": "infeasible",
                "summary": {},
                "hourly": [],
            })
            continue

        mc.update_after_optimize(res, load_kw=load)
        soc_now = res["soc"][-1]

        # 组装逐时明细
        hourly = []
        for t in range(24):
            hourly.append({
                "h": t,
                "load_kW": round(float(load[t]), 1),
                "pv_kW": round(float(pv[t]), 1),
                "pch_kW": round(float(res["pch"][t]), 1),
                "pdis_kW": round(float(res["pdis"][t]), 1),
                "pgrid_kW": round(float(res["pgrid"][t]), 1),
                "flex_kW": round(float(res["flex"][t]), 1),
                "flex_down_kW": round(float(res["flex_down"][t]), 1),
                "soc": round(float(res["soc"][t]), 4),
            })

        summary = {
            "load_kWh": round(float(load.sum()), 0),
            "pv_kWh": round(float(pv.sum()), 0),
            "buy_kWh": round(float(res["buy_kwh"]), 0),
            "export_kWh": round(float(res["export_kwh"]), 0),
            "net_grid_kWh": round(float(res["net_grid_kwh"]), 0),
            "green_ratio_pct": round(float(res["green_ratio"]), 2),
            "cost_yuan": round(float(res["cost_yuan"]), 1),
            "chg_kWh": round(float(res["chg_kwh"]), 0),
            "dis_kWh": round(float(res["dis_kwh"]), 0),
            "chg_pv_kWh": round(float(res["chg_pv_kwh"]), 0),
            "dis_replace_kWh": round(float(res["dis_replace_kwh"]), 0),
        }

        days.append({
            "date": _date_str(start_date, d),
            "idx": d,
            "mode": mode,
            "is_shock": bool(is_shock),
            "flex_min": round(flex_min_override, 3),
            "status": "ok",
            "summary": summary,
            "hourly": hourly,
        })

    meta = {
        "total_days": day_num,
        "fail_days": fail_days,
        "start_date": _date_str(start_date, 0),
        "end_date": _date_str(start_date, day_num - 1),
        "battery": {
            "E_max_kWh": config.E_BAT_MAX,
            "P_max_kW": config.P_BAT_MAX,
            "eta_ch": config.ETA_CH,
            "eta_dis": config.ETA_DIS,
            "soc_range": [config.SOC_MIN, config.SOC_MAX],
            "soc_init": config.SOC_INIT,
        },
        "pv_scale": config.PV_SCALE,
        "flex": {
            "enabled": config.FLEX_ENABLED,
            "flex_min": config.FLEX_MIN,
            "flex_min_safe": config.FLEX_MIN_SAFE,
            "flex_energy_ratio": config.FLEX_ENERGY_RATIO,
        },
        "grid_export": config.ALLOW_GRID_EXPORT,
        "weights": {"w_cost": config.W_COST, "w_green": config.W_GREEN},
    }

    return {"meta": meta, "days": days}


def _date_str(start_date: str, d: int) -> str:
    """由起始日期 + 天偏移生成 YYYY-MM-DD。"""
    from datetime import datetime, timedelta
    base = datetime.strptime(start_date, "%Y-%m-%d")
    return (base + timedelta(days=d)).strftime("%Y-%m-%d")


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        config.BASE_DIR, "output", "yearly_dispatch.json")

    print("开始逐日调度导出（储能 %.0f MWh / %.0f MW）..." %
          (config.E_BAT_MAX / 1000, config.P_BAT_MAX / 1000))
    result = build_yearly_dispatch()
    days = result["days"]
    meta = result["meta"]

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)

    # 简要校验
    ok_days = [d for d in days if d["status"] == "ok"]
    print(f"完成：共 {meta['total_days']} 天，无解 {meta['fail_days']} 天")
    print(f"日期范围：{meta['start_date']} ~ {meta['end_date']}")
    print(f"JSON 已保存：{out_path}")

    # 抽查首尾两天
    for probe in [ok_days[0], ok_days[-1]]:
        s = probe["summary"]
        print(f"\n示例 {probe['date']} ({probe['mode']}模式): "
              f"负荷 {s['load_kWh']:.0f} kWh, 光伏 {s['pv_kWh']:.0f} kWh, "
              f"购电 {s['buy_kWh']:.0f} kWh, 余电 {s['export_kWh']:.0f} kWh, "
              f"绿电 {s['green_ratio_pct']:.1f}%, 成本 {s['cost_yuan']:.0f} 元")
        print(f"  储能: 充电 {s['chg_kWh']:.0f} kWh, 放电 {s['dis_kWh']:.0f} kWh, "
              f"充电来自光伏 {s['chg_pv_kWh']:.0f} kWh")

    return out_path


if __name__ == "__main__":
    main()
