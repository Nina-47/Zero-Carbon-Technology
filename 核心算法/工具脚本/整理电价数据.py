# -*- coding: utf-8 -*-
"""
整理电价数据 → 与负荷分表同构的「说明 / 逐时数据 / 日汇总」三 sheet 格式。

输入：电力天气整合数据_逐小时.xlsx（根目录，577 天逐时）
      关键两列：节点日前电价_元MWh、节点实时电价_元MWh

输出：02_原始数据/电价数据/广东电力市场电价_逐时.xlsx
  - 说明：电价口径（日前 vs 实时）、单位、来源、范围
  - 逐时数据：日期 | 节点日前电价(24h) | 节点实时电价(24h)，两行表头
  - 日汇总：日期 | 日前最高/最低/均价 + 实时最高/最低/均价

单位保持「元/MWh」与模块二 config.load_realtime_price 对齐（其内部 ÷1000 转元/kWh）。
"""
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:/Users/xiaoY/Desktop/零碳科技"
SRC = f"{BASE}/电力天气整合数据_逐小时.xlsx"
OUT_DIR = f"{BASE}/02_原始数据/电价数据"
OUT = f"{OUT_DIR}/广东电力市场电价_逐时.xlsx"


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return datetime.date.fromisoformat(s[:10])
    return None


# ---------- 读原文件 ----------
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb.active
rows = ws.iter_rows(values_only=True)
header = next(rows)
idx = {h: i for i, h in enumerate(header)}
print("表头:", header)

days = {}
for r in rows:
    d = parse_date(r[0])
    if d is None:
        continue
    h = int(r[idx["小时"]])
    da = r[idx["节点日前电价_元MWh"]]
    rt = r[idx["节点实时电价_元MWh"]]
    days.setdefault(d, {"日前": {}, "实时": {}})
    days[d]["日前"][h] = float(da) if isinstance(da, (int, float)) else None
    days[d]["实时"][h] = float(rt) if isinstance(rt, (int, float)) else None
wb.close()

dates = sorted(days.keys())
print(f"整理到 {len(dates)} 天，范围 {dates[0]} ~ {dates[-1]}")

HOURS_CN = [f"{h}时" for h in range(24)]

# ---------- 样式 ----------
TYPE_FILL = PatternFill("solid", fgColor="BDD7EE")
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _fill(ws, row, cols, style_fill, style_font=BOLD):
    for c in cols:
        ws.cell(row, c).fill = style_fill
        ws.cell(row, c).font = style_font
        ws.cell(row, c).alignment = CENTER


out_wb = openpyxl.Workbook()

# ---- Sheet1 说明 ----
ws0 = out_wb.active
ws0.title = "说明"
lines = [
    ["广东电力市场 电价数据（节点日前 + 节点实时）"],
    [""],
    ["两列电价含义："],
    ["  节点日前电价 = 电力现货市场「日前市场」出清价（提前一天成交的价格）"],
    ["  节点实时电价 = 电力现货市场「实时市场」出清价（当天实际运行价格）"],
    [""],
    ["单位：元/MWh（兆瓦时电价）。模块二调度里会 ÷1000 转成 元/kWh 使用。"],
    [""],
    ["数据范围：2025-01-01 ~ 2026-07-31（577 天，逐时完整无缺失）"],
    ["数据来源：电力天气整合数据_逐小时.xlsx（原文件另含统调负荷与天气字段，已剥离）"],
    [""],
    ["说明：夜间（0-6时）日前/实时电价可能为 0 或负值，属现货市场真实现象（低谷时"],
    ["      电力供大于求），调度代码中已对负价做 clip 处理。"],
]
for row in lines:
    ws0.append(row)
ws0.column_dimensions["A"].width = 70

# ---- Sheet2 逐时数据 ----
ws1 = out_wb.create_sheet("逐时数据")
ws1.cell(1, 1, "日期")
ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
ws1.cell(1, 2, "节点日前电价(元/MWh)")
ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=25)
ws1.cell(1, 26, "节点实时电价(元/MWh)")
ws1.merge_cells(start_row=1, start_column=26, end_row=1, end_column=49)
for i, h in enumerate(HOURS_CN):
    ws1.cell(2, 2 + i, h)
    ws1.cell(2, 26 + i, h)
_fill(ws1, 1, range(1, 50), TYPE_FILL)
_fill(ws1, 2, range(1, 50), HEADER_FILL)
for d in dates:
    row = [d.strftime("%Y-%m-%d")]
    for key in ("日前", "实时"):
        for h in range(24):
            v = days[d][key].get(h)
            row.append(round(v, 2) if v is not None else None)
    ws1.append(row)
ws1.freeze_panes = "B3"

# ---- Sheet3 日汇总 ----
ws2 = out_wb.create_sheet("日汇总")
cols = ["日期", "日前最高", "日前最低", "日前均价", "实时最高", "实时最低", "实时均价"]
ws2.append(cols)
_fill(ws2, 1, range(1, 8), HEADER_FILL)
for d in dates:
    row = [d.strftime("%Y-%m-%d")]
    for key in ("日前", "实时"):
        vals = [v for v in days[d][key].values() if v is not None]
        if vals:
            row += [round(max(vals), 2), round(min(vals), 2), round(sum(vals) / len(vals), 2)]
        else:
            row += [None, None, None]
    ws2.append(row)
for c in range(1, 8):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12
ws2.freeze_panes = "A2"

import os
os.makedirs(OUT_DIR, exist_ok=True)
out_wb.save(OUT)
print(f"\n已保存: {OUT}")
print(f"  逐时数据: {len(dates)} 天 × 48 列（日前24 + 实时24）")
print(f"  日汇总: {len(dates)} 天 × 7 列")
