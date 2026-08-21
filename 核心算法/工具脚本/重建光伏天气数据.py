# -*- coding: utf-8 -*-
"""
重建「光伏 × 天气」训练数据（合并数据集_光伏×天气.csv）

用途：喂给 优化调度/pv_forecast.py 做光伏预测（KNN 相似日 + 两步法）。

列格式（对齐 pv_forecast.py 的 _get_hourly 读取约定，h 用「时」后缀）：
  日期
  发电_kWh_0时 ~ 发电_kWh_23时      （一期+二期+车棚 逐时发电量 kWh，数值=平均功率 kW）
  辐射_MJpm2_0时 ~ 辐射_MJpm2_23时  （逐时太阳辐射，MJ/m²）
  气温_℃_0时 ~ 气温_℃_23时          （逐时气温 ℃）
  日总辐射_MJpm2                    （逐时辐射求和）
  季节                              （春/夏/秋/冬）

输入：
  - 02_原始数据/光伏数据/中嘉光伏数据.xlsx        （一期/二期/车棚 逐时 kWh）
  - 02_原始数据/天气数据/中山天气_OpenMeteo.xlsx  （逐时宽表，辐射 W/m² → 转 MJ/m²）

输出：
  - 04_预测输出/V4_优化调度/合并数据集_光伏×天气.csv
"""
import sys
import os
import datetime
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:/Users/xiaoY/Desktop/零碳科技"
PV_FILE = f"{BASE}/02_原始数据/光伏数据/中嘉光伏数据.xlsx"
WX_FILE = f"{BASE}/02_原始数据/天气数据/中山天气_OpenMeteo.xlsx"
OUT_DIR = f"{BASE}/04_预测输出/V4_优化调度"
OUT = f"{OUT_DIR}/合并数据集_光伏×天气.csv"

W2MJ = 0.0036  # W/m²(小时平均) → MJ/m²(小时累积)：×3600s ÷ 1e6


def load_pv_hourly(sheet_name):
    """读光伏逐时 kWh，返回 {date_str: [24h]}。列：r[3]=对应时间, r[7:31]=0点~23点"""
    wb = openpyxl.load_workbook(PV_FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]
    out = {}
    for r in ws.iter_rows(values_only=True):
        t = r[3]
        if t is None:
            continue
        if isinstance(t, datetime.datetime):
            d = t.date()
        elif isinstance(t, str):
            try:
                d = datetime.date.fromisoformat(t.strip()[:10])
            except ValueError:
                continue
        else:
            continue
        vals = [float(x) if isinstance(x, (int, float)) else 0.0 for x in r[7:31]]
        out[d.isoformat()] = (vals + [0.0] * 24)[:24]
    wb.close()
    return out


def load_openmeteo():
    """读 OpenMeteo 逐时宽表，返回 {date_str: dict(辐射[24]MJ, 气温[24]℃)}"""
    wb = openpyxl.load_workbook(WX_FILE, read_only=True, data_only=True)
    ws = wb["逐时宽表"]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr) if h is not None}
    rad_cols = [idx[f"辐射_Wm2_{h:02d}h"] for h in range(24)]
    temp_cols = [idx[f"气温_℃_{h:02d}h"] for h in range(24)]
    out = {}
    for r in rows:
        if r[0] is None or not isinstance(r[0], str):
            continue
        d = r[0].strip()[:10]
        rad = [float(r[c]) * W2MJ if isinstance(r[c], (int, float)) else 0.0 for c in rad_cols]
        temp = [float(r[c]) if isinstance(r[c], (int, float)) else 0.0 for c in temp_cols]
        out[d] = {"rad": rad, "temp": temp}
    wb.close()
    return out


def season_of(d):
    m = d.month
    if m in (3, 4, 5):
        return "春"
    if m in (6, 7, 8):
        return "夏"
    if m in (9, 10, 11):
        return "秋"
    return "冬"


# ---------- 读取 ----------
pv = {"一期": load_pv_hourly("中嘉污水处理厂8MWp光伏电站(一期)"),
      "二期": load_pv_hourly("中嘉污水处理厂8MWp光伏电站(二期)"),
      "车棚": load_pv_hourly("中嘉污水处理厂光伏车棚项目")}
wx = load_openmeteo()

# 日期交集
dates = sorted(set.intersection(*[set(pv[k]) for k in pv], set(wx)))
dates = [datetime.date.fromisoformat(d) for d in dates]
print(f"光伏: 一期={len(pv['一期'])} 二期={len(pv['二期'])} 车棚={len(pv['车棚'])} 天")
print(f"天气: {len(wx)} 天")
print(f"交集日期: {len(dates)} 天 ({dates[0]} ~ {dates[-1]})")

# ---------- 组装 ----------
rows = []
for d in dates:
    ds = d.isoformat()
    pv24 = [pv["一期"][ds][h] + pv["二期"][ds][h] + pv["车棚"][ds][h] for h in range(24)]
    rad24 = wx[ds]["rad"]
    temp24 = wx[ds]["temp"]
    rad_daily = round(sum(rad24), 3)
    row = [ds]
    row += [round(v, 2) for v in pv24]
    row += [round(v, 3) for v in rad24]
    row += [round(v, 1) for v in temp24]
    row += [rad_daily, season_of(d)]
    rows.append(row)

cols = ["日期"]
cols += [f"发电_kWh_{h}时" for h in range(24)]
cols += [f"辐射_MJpm2_{h}时" for h in range(24)]
cols += [f"气温_℃_{h}时" for h in range(24)]
cols += ["日总辐射_MJpm2", "季节"]

df = pd.DataFrame(rows, columns=cols)
os.makedirs(OUT_DIR, exist_ok=True)
df.to_csv(OUT, index=False, encoding="utf-8-sig")
print(f"\n已保存: {OUT}")
print(f"  共 {len(df)} 天，{len(cols)} 列")
print(f"  示例 {df['日期'].iloc[0]}: 光伏日总 {sum(rows[0][1:25]):.0f} kWh, "
      f"辐射日总 {df['日总辐射_MJpm2'].iloc[0]} MJ/m², 季节 {df['季节'].iloc[0]}")
