# -*- coding: utf-8 -*-
"""
爬取中山市逐时天气数据（Open-Meteo），存成项目天气库。

数据源：Open-Meteo（免费、无 key、历史 + 预报一体）
  - archive  API：历史实测（通常延迟约 5 天）
  - forecast API：最近回看 + 未来预报

坐标：中山市 22.52°N, 113.39°E

字段（逐时）：
  temperature_2m        气温 ℃
  relative_humidity_2m  相对湿度 %
  precipitation         降水量 mm
  shortwave_radiation   短波辐射 W/m²（瞬时）
  cloud_cover           云量 %
  wind_speed_10m        10m 风速 km/h

输出：02_原始数据/天气数据/中山天气_OpenMeteo.xlsx
  - 逐时宽表：日期 + 6 要素 × 24 小时
  - 逐日摘要：最高温/最低温/日降水/日辐射(MJ)/均湿度/均云量
  - 长表：datetime 逐行（前端友好）
"""
import sys, io, json, datetime, urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding='utf-8')

LAT, LON = 22.52, 113.39
TZ = "Asia/Shanghai"
FIELDS = ("temperature_2m,relative_humidity_2m,precipitation,"
          "shortwave_radiation,cloud_cover,wind_speed_10m")
FIELD_NAMES = ["气温_℃", "湿度_%", "降水_mm", "辐射_Wm2", "云量_%", "风速_kmh"]

OUT = r"C:/Users/xiaoY/Desktop/零碳科技/02_原始数据/天气数据/中山天气_OpenMeteo.xlsx"

TODAY = datetime.date.today()  # 2026-08-21


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=90) as r:
        return json.loads(r.read().decode("utf-8"))


def fetch_archive(start: str, end: str):
    url = (f"https://archive-api.open-meteo.com/v1/archive"
           f"?latitude={LAT}&longitude={LON}"
           f"&start_date={start}&end_date={end}"
           f"&hourly={FIELDS}&timezone={TZ}")
    return fetch(url)


def fetch_forecast(past_days: int = 8, forecast_days: int = 14):
    url = (f"https://api.open-meteo.com/v1/forecast"
           f"?latitude={LAT}&longitude={LON}"
           f"&hourly={FIELDS}&timezone={TZ}"
           f"&past_days={past_days}&forecast_days={forecast_days}")
    return fetch(url)


# ---------- 1. 历史（archive，到最新可用，约延迟5天） ----------
archive_end = (TODAY - datetime.timedelta(days=6)).isoformat()
print(f"爬取历史: 2025-07-01 ~ {archive_end} (archive)...")
data_hist = fetch_archive("2025-07-01", archive_end)

# ---------- 2. 最近几天 + 未来预报（forecast，past_days 回看 + forecast_days 预报） ----------
print(f"爬取最近+预报: 回看8天 + 未来14天 (forecast)...")
data_fc = fetch_forecast(past_days=8, forecast_days=14)


def to_records(data, source):
    """把 Open-Meteo 返回的 hourly JSON 转成 [(datetime, [6值], source)]"""
    h = data.get("hourly", {})
    times = h.get("time", [])
    keys = ["temperature_2m", "relative_humidity_2m", "precipitation",
            "shortwave_radiation", "cloud_cover", "wind_speed_10m"]
    recs = []
    for i, t in enumerate(times):
        dt = datetime.datetime.fromisoformat(t)
        vals = [h[k][i] if i < len(h.get(k, [])) else None for k in keys]
        recs.append((dt, vals, source))
    return recs


recs = to_records(data_hist, "archive") + to_records(data_fc, "forecast")
# 按时间去重（forecast 与 archive 重叠部分以 archive 实测为准）
seen = {}
for dt, vals, src in recs:
    if dt not in seen or src == "archive":
        seen[dt] = (dt, vals, src)
merged = sorted(seen.values(), key=lambda x: x[0])
print(f"合并后总逐时记录: {len(merged)} 条，"
      f"范围 {merged[0][0]} ~ {merged[-1][0]}")

# 按天分组
days = {}
for dt, vals, src in merged:
    d = dt.date()
    h = dt.hour
    days.setdefault(d, {})[h] = vals

# 只保留完整24小时的日期
full_days = {d: v for d, v in days.items() if len(v) == 24}
dates = sorted(full_days.keys())
print(f"完整日期: {len(dates)} 天 ({dates[0]} ~ {dates[-1]})")

# ---------- 写 Excel ----------
wb = openpyxl.Workbook()
HDR_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

# Sheet1 逐时宽表
ws1 = wb.active
ws1.title = "逐时宽表"
ws1.append(["日期"] + [f"{n}_{h:02d}h" for n in FIELD_NAMES for h in range(24)])
for c in range(1, 1 + 1 + 6 * 24):
    ws1.cell(1, c).fill = HDR_FILL
    ws1.cell(1, c).font = BOLD
    ws1.cell(1, c).alignment = CENTER
for d in dates:
    row = [d.strftime("%Y-%m-%d")]
    for h in range(24):
        vals = full_days[d][h]
        row.extend(vals)
    ws1.append(row)
ws1.freeze_panes = "B2"

# Sheet2 逐日摘要
ws2 = wb.create_sheet("逐日摘要")
ws2.append(["日期", "最高温℃", "最低温℃", "日降水mm", "日辐射MJ/m²",
            "均湿度%", "均云量%", "均风速km/h"])
for c in range(1, 9):
    ws2.cell(1, c).fill = HDR_FILL
    ws2.cell(1, c).font = BOLD
    ws2.cell(1, c).alignment = CENTER
for d in dates:
    temps = [full_days[d][h][0] for h in range(24)]
    humids = [full_days[d][h][1] for h in range(24)]
    precs = [full_days[d][h][2] for h in range(24)]
    rads = [full_days[d][h][3] for h in range(24)]
    clouds = [full_days[d][h][4] for h in range(24)]
    winds = [full_days[d][h][5] for h in range(24)]

    def _mean(vs):
        vs = [v for v in vs if v is not None]
        return round(sum(vs) / len(vs), 2) if vs else None

    rad_sum_mj = round(sum(v for v in rads if v is not None) * 0.0036, 2)  # W/m²×h → MJ/m²
    ws2.append([
        d.strftime("%Y-%m-%d"),
        round(max(temps), 1), round(min(temps), 1),
        round(sum(v for v in precs if v is not None), 2),
        rad_sum_mj, _mean(humids), _mean(clouds), _mean(winds),
    ])
ws2.freeze_panes = "A2"

# Sheet3 长表（前端友好）
ws3 = wb.create_sheet("长表")
ws3.append(["datetime", "date", "hour", *FIELD_NAMES])
for c in range(1, 1 + 9):
    ws3.cell(1, c).fill = HDR_FILL
    ws3.cell(1, c).font = BOLD
    ws3.cell(1, c).alignment = CENTER
for dt, vals, src in merged:
    d = dt.date()
    if d not in full_days:
        continue
    ws3.append([dt.strftime("%Y-%m-%d %H:%M"), d.strftime("%Y-%m-%d"),
                dt.hour, *vals])
ws3.freeze_panes = "A2"

wb.save(OUT)
print(f"\n已保存: {OUT}")
print(f"  逐时宽表: {len(dates)} 天")
print(f"  覆盖: {dates[0]} ~ {dates[-1]}")
