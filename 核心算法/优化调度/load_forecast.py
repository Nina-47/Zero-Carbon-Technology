# -*- coding: utf-8 -*-
"""
负荷预测模块：天气 → 逐时负荷（三厂合计示范流程）

数据：合并数据集_负荷×天气.xlsx（三厂 A/B/C 合计负荷，395天）
方法：KNN 相似日匹配——按"星期 + 季节 + 日均温"找最相似的历史日，
     用其逐时负荷形状加权合成目标日预测。

定位：示范"天气/日历 → 负荷预测 → 调度"的完整链路。
     当前调度模型用 data_source 的单厂负荷，本模块用三厂合计口径，
     属口径未对齐的示范实现；后续对齐单厂数据后替换训练源即可。

注：与 pv_forecast.py 同构的 KNN 思路，方便答辩统一解释方法。
"""

import os
import numpy as np
import openpyxl

import config


def load_load_weather():
    """读负荷×天气数据，返回 dict {dates, hourly_load(24, n×24), dow, weekend, temp_mean, season}。"""
    path = os.path.join(config.PROJECT_ROOT, "5-过程分析数据", "合并数据集_负荷×天气.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}
    load_cols = [idx[f"负荷_{h}h"] for h in range(24)]
    temp_cols = [idx[f"气温_{h}h"] for h in range(24)]

    dates, loads, dows, weekend, temp_mean, seasons = [], [], [], [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        lv = [float(row[c]) if row[c] is not None else 0.0 for c in load_cols]
        tv = [float(row[c]) if row[c] is not None else 0.0 for c in temp_cols]
        dates.append(str(row[0])[:10])
        loads.append(lv)
        dows.append(row[idx["星期"]] if row[idx["星期"]] is not None else 0)
        weekend.append(row[idx["是否周末"]] if row[idx["是否周末"]] is not None else 0)
        temp_mean.append(float(np.mean(tv)))
        seasons.append(str(row[idx["季节"]]) if row[idx["季节"]] is not None else "")

    return {
        "dates": dates,
        "hourly_load": np.array(loads),       # (n, 24) MW
        "dow": np.array(dows),
        "weekend": np.array(weekend),
        "temp_mean": np.array(temp_mean),
        "season": seasons,
    }


def forecast_next_day(data, k=5):
    """
    预测"明天"的逐时负荷（三厂合计，MW）。
    在历史库里按 星期/季节/日均温 相似度匹配 K 个相似日，加权合成。

    返回 (hourly_pred(24, MW), matched_dates) 或 (None, [])。
    """
    n = len(data["dates"])
    if n < k:
        return None, []

    # 目标日特征：默认预测"下一天"。温度用近 7 日均温外推（简化）。
    target_dow = (data["dow"][-1] + 1) % 7
    target_weekend = 1 if target_dow >= 5 else 0
    target_temp = float(np.mean(data["temp_mean"][-7:]))

    dists = []
    for i in range(n):
        dow_dist = 0.0 if data["dow"][i] == target_dow else 1.0
        weekend_dist = 0.0 if data["weekend"][i] == target_weekend else 0.3
        temp_dist = abs(data["temp_mean"][i] - target_temp) / (target_temp + 1e-6)
        dist = 0.5 * dow_dist + 0.2 * weekend_dist + 0.3 * temp_dist
        dists.append((dist, i))

    dists.sort(key=lambda x: x[0])
    top_k = dists[:min(k, n)]
    weights = np.array([1.0 / (d + 0.01) for d, _ in top_k])
    weights /= weights.sum()
    profiles = np.array([data["hourly_load"][i] for _, i in top_k])
    hourly_pred = np.average(profiles, axis=0, weights=weights)
    matched_dates = [data["dates"][i] for _, i in top_k]
    return hourly_pred, matched_dates


if __name__ == "__main__":
    data = load_load_weather()
    pred, matched = forecast_next_day(data, k=5)
    if pred is not None:
        print(f"负荷预测(三厂合计): 日总 {pred.sum():.0f} MW，峰值 {pred.max():.1f} MW")
        print(f"匹配相似日: {matched}")
